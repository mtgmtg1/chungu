#!/usr/bin/env python3
# [Flow: Step 1 (업로드 -> 파일 유형 감지/압축 해제/Storage 저장) -> Step 2 (비용 계산) -> Step 3 (승인 -> 포인트 차감 + Celery) -> Step 4 (상태 폴링/Storage 다운로드)]
import json
import tempfile
import uuid
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import List

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import select
from sqlalchemy.orm import Session

from .. import settings_store
from ..auth.supabase_auth import CurrentUser, get_current_admin, get_current_user
from ..core import archive_handler, media_loader, points_service, supabase_client
from ..core.prompts import DEFAULT_COLUMNS
from ..db.models import Job
from ..db.session import get_db
from ..workers.tasks import run_job

router = APIRouter(prefix="/api", tags=["jobs"])


MEDIA_EXTENSIONS = {
    ".pdf", ".zip", ".rar", ".7z", ".tar", ".gz", ".tgz", ".bz2",
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tiff", ".tif",
    ".mp3", ".wav", ".flac", ".aac", ".ogg", ".m4a", ".wma",
    ".mp4", ".avi", ".mov", ".mkv", ".flv", ".wmv", ".webm", ".m4v",
}


def _parse_columns(raw: str) -> list[str]:
    raw = (raw or "").strip()
    if not raw:
        return list(DEFAULT_COLUMNS)
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list) and parsed:
            return [str(c).strip() for c in parsed if str(c).strip()]
    except json.JSONDecodeError:
        pass
    return [c.strip() for c in raw.split(",") if c.strip()]


@router.post("/jobs/upload")
async def upload_job(
    files: List[UploadFile] = File(...),
    pipeline: str = Form("vision"),
    columns: str = Form(""),
    prompt: str = Form(""),
    dpi: int = Form(150),
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not files:
        raise HTTPException(status_code=400, detail="파일을 선택하세요")
    if pipeline not in ("vision", "hybrid"):
        pipeline = settings_store.get_setting(db, "default_pipeline") or "vision"

    max_mb = int(settings_store.get_setting(db, "max_file_mb") or "200")
    total_size = 0
    file_data: List[bytes] = []
    for file in files:
        if not file.filename:
            raise HTTPException(status_code=400, detail="파일 이름이 없습니다")
        ext = Path(file.filename).suffix.lower()
        if ext not in MEDIA_EXTENSIONS:
            raise HTTPException(status_code=400, detail=f"지원하지 않는 파일 형식입니다: {file.filename}")
        data = await file.read()
        total_size += len(data)
        file_data.append(data)

    if total_size > max_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail=f"전체 파일이 너무 큽니다 (최대 {max_mb}MB)")

    # 하나의 파일만 업로드되고 PDF일 때는 기존 단일 PDF 플로우를 유지
    is_single_pdf = len(files) == 1 and files[0].filename.lower().endswith(".pdf")
    original_filename = files[0].filename if len(files) == 1 else f"{len(files)}_files"

    job = Job(
        user_id=uuid.UUID(user.user_id),
        email=user.email,
        pipeline=pipeline,
        columns=_parse_columns(columns),
        prompt=prompt.strip(),
        dpi=dpi,
        original_filename=original_filename,
        status="pending",
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    pages = 0
    image_count = 0
    audio_seconds = 0
    video_seconds = 0
    total_files = 0
    file_type = "pdf"

    try:
        if is_single_pdf:
            data = file_data[0]
            pages = len(PdfReader(BytesIO(data)).pages)
            total_files = 1
            storage_path = supabase_client.upload_pdf(job.id, data, files[0].filename)
            job.pdf_storage_path = storage_path
            job.file_type = "pdf"
        else:
            # 여러 파일/아카이브: 임시 디렉터리에 저장 후 분석
            with tempfile.TemporaryDirectory() as tmpdir:
                tmp_path = Path(tmpdir)
                extracted: list[Path] = []
                for file, data in zip(files, file_data):
                    if archive_handler.is_archive(file.filename):
                        archive_dest = tmp_path / f"extracted_{file.filename}"
                        archive_dest.mkdir(parents=True, exist_ok=True)
                        extracted.extend(archive_handler.extract_all_recursive(file.filename, data, archive_dest))
                    else:
                        file_path = tmp_path / file.filename
                        file_path.parent.mkdir(parents=True, exist_ok=True)
                        file_path.write_bytes(data)
                        extracted.append(file_path)

                for fp in extracted:
                    ftype = media_loader.detect_file_type(fp)
                    if ftype == "pdf":
                        try:
                            pages += len(PdfReader(fp).pages)
                        except Exception:
                            pass
                        total_files += 1
                    elif ftype == "image":
                        image_count += 1
                        total_files += 1
                    elif ftype == "audio":
                        audio_seconds += media_loader.get_media_duration_seconds(fp)
                        total_files += 1
                    elif ftype == "video":
                        video_seconds += media_loader.get_media_duration_seconds(fp)
                        total_files += 1

                job.total_files = total_files
                job.media_duration_seconds = audio_seconds + video_seconds
                job.extracted_files = [
                    {"path": str(p.relative_to(tmp_path)), "type": media_loader.detect_file_type(p), "size": p.stat().st_size}
                    for p in extracted
                ]

                # Storage에는 원본 파일들을 압축하여 하나로 업로드
                if len(files) == 1:
                    storage_path = supabase_client.upload_input(job.id, file_data[0], files[0].filename)
                    job.pdf_storage_path = storage_path
                    job.file_type = "archive" if archive_handler.is_archive(files[0].filename) else "mixed"
                else:
                    zip_path = tmp_path / f"{job.id}.zip"
                    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                        for file, data in zip(files, file_data):
                            zf.writestr(file.filename, data)
                    storage_path = supabase_client.upload_input(job.id, zip_path.read_bytes(), zip_path.name, "application/zip")
                    job.pdf_storage_path = storage_path
                    job.file_type = "mixed"

        max_pages = int(settings_store.get_setting(db, "max_pages") or "2000")
        if pages > max_pages:
            db.delete(job)
            db.commit()
            raise HTTPException(status_code=413, detail=f"페이지가 너무 많습니다 (최대 {max_pages})")

        job.total_pages = pages
        db.commit()
    except HTTPException:
        raise
    except Exception as e:
        db.delete(job)
        db.commit()
        raise HTTPException(status_code=502, detail=f"파일 처리 실패: {e}")

    cost = points_service.calculate_cost(db, pages=pages, image_count=image_count, audio_seconds=audio_seconds, video_seconds=video_seconds)
    return {
        "job_id": job.id,
        "status": job.status,
        "file_type": job.file_type,
        "total_pages": pages,
        "total_files": total_files,
        "media_duration_seconds": audio_seconds + video_seconds,
        "cost": cost,
        "balance": user.points_balance,
    }


@router.put("/jobs/{job_id}")
def update_job(
    job_id: str,
    payload: dict,
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    job = db.get(Job, job_id)
    if job is None or str(job.user_id) != user.user_id:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")
    if job.status != "pending":
        raise HTTPException(status_code=400, detail="대기 중인 작업만 수정할 수 있습니다")

    if "pipeline" in payload and payload["pipeline"] in ("vision", "hybrid"):
        job.pipeline = payload["pipeline"]
    if "columns" in payload:
        job.columns = _parse_columns(payload["columns"])
    if "prompt" in payload:
        job.prompt = str(payload["prompt"]).strip()
    db.commit()
    return _job_summary(job)


@router.post("/jobs/{job_id}/confirm")
def confirm_job(
    job_id: str,
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    job = db.get(Job, job_id)
    if job is None or str(job.user_id) != user.user_id:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")
    if job.status != "pending":
        raise HTTPException(status_code=400, detail="이미 처리되었거나 취소된 작업입니다")

    from ..db.models import User

    db_user = db.get(User, job.user_id)
    if db_user is None:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")

    # 작업 정보에서 비용 재계산
    pages = job.total_pages
    image_count = 0
    audio_seconds = 0
    video_seconds = 0
    for info in job.extracted_files or []:
        ftype = info.get("type", "")
        if ftype == "image":
            image_count += 1
        elif ftype == "audio":
            audio_seconds += info.get("duration", 0)
        elif ftype == "video":
            video_seconds += info.get("duration", 0)
    if job.file_type == "pdf":
        image_count = 0
        audio_seconds = 0
        video_seconds = 0

    cost = points_service.calculate_cost(db, pages=pages, image_count=image_count, audio_seconds=audio_seconds, video_seconds=video_seconds)
    try:
        points_service.spend_points(db, db_user, cost["points"], f"미디어 작업: {job.original_filename}")
    except ValueError as e:
        raise HTTPException(status_code=402, detail=str(e))

    job.cost_points = cost["points"]
    job.status = "queued"
    db.commit()

    run_job.delay(job.id)
    return {"job_id": job.id, "status": job.status, "remaining_points": db_user.points_balance}


@router.get("/jobs")
def list_jobs(
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
    limit: int = 100,
):
    rows = db.execute(
        select(Job).where(Job.user_id == uuid.UUID(user.user_id)).order_by(Job.created_at.desc()).limit(limit)
    ).scalars().all()
    return [_job_summary(j) for j in rows]


@router.get("/jobs/{job_id}")
def get_job(job_id: str, user: CurrentUser = Depends(get_current_user), db: Session = Depends(get_db)):
    job = db.get(Job, job_id)
    if job is None or str(job.user_id) != user.user_id:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")
    return _job_summary(job)


@router.get("/jobs/{job_id}/download")
def download_job(
    job_id: str,
    type: str = "xlsx",
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    job = db.get(Job, job_id)
    if job is None or str(job.user_id) != user.user_id:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")
    if job.status != "done":
        raise HTTPException(status_code=400, detail="완료된 작업만 다운로드할 수 있습니다")

    path_map = {
        "csv": job.result_csv_storage_path,
        "md": job.result_md_storage_path,
        "xlsx": job.result_xlsx_storage_path,
    }
    path = path_map.get(type)
    if not path:
        raise HTTPException(status_code=404, detail="결과 파일이 없습니다")

    try:
        url = supabase_client.get_signed_download_url(path, bucket="results", expires_in=3600)
        return {"download_url": url}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f"다운로드 링크 생성 실패: {e}")


@router.get("/jobs/{job_id}/preview")
def preview_job(
    job_id: str,
    user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    job = db.get(Job, job_id)
    if job is None or str(job.user_id) != user.user_id:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")
    if not job.result_xlsx_storage_path:
        raise HTTPException(status_code=400, detail="결과 파일이 준비되지 않았습니다")

    try:
        xlsx_bytes = supabase_client.get_service_client().storage.from_("results").download(job.result_xlsx_storage_path)
        wb = load_workbook(BytesIO(xlsx_bytes))
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheets[sheet_name] = [[cell.value for cell in row] for row in ws.iter_rows()]
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"결과 미리보기 생성 실패: {e}")

    source_url = None
    if job.pdf_storage_path:
        try:
            source_url = supabase_client.get_signed_download_url(job.pdf_storage_path, bucket="pdfs", expires_in=3600)
        except Exception:
            pass

    return {
        "job": _job_summary(job),
        "sheets": sheets,
        "source_url": source_url,
    }


@router.get("/admin/jobs")
def admin_list_jobs(
    admin: CurrentUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
    limit: int = 100,
):
    rows = db.execute(select(Job).order_by(Job.created_at.desc()).limit(limit)).scalars().all()
    return [_job_summary(j) for j in rows]


def _job_summary(job: Job) -> dict:
    return {
        "job_id": job.id,
        "status": job.status,
        "pipeline": job.pipeline,
        "file_type": job.file_type,
        "filename": job.original_filename,
        "total_pages": job.total_pages,
        "done_pages": job.done_pages,
        "total_files": job.total_files,
        "done_files": job.done_files,
        "media_duration_seconds": job.media_duration_seconds,
        "cost_points": job.cost_points,
        "error_log": job.error_log,
        "created_at": job.created_at.isoformat() if job.created_at else None,
        "finished_at": job.finished_at.isoformat() if job.finished_at else None,
        "downloadable": job.status == "done",
    }


# 하위 호환: 기존 /download/{token} 엔드포인트는 Storage URL로 리디렉션
@router.get("/download/{token}")
def legacy_download(token: str, type: str = "csv", db: Session = Depends(get_db)):
    job = db.get(Job, token)
    if job is None or job.download_token != token:
        raise HTTPException(status_code=404, detail="유효하지 않은 다운로드 링크입니다")
    if job.expires_at and job.expires_at < datetime.utcnow():
        raise HTTPException(status_code=410, detail="다운로드 링크가 만료되었습니다")

    path = job.result_csv_storage_path if type == "csv" else job.result_md_storage_path
    if not path:
        # Storage로 이전되지 않은 예전 파일은 로컬 경로 사용
        local = job.result_csv_path if type == "csv" else job.result_md_path
        if not local or not Path(local).exists():
            raise HTTPException(status_code=404, detail="결과 파일이 없습니다")
        base = Path(job.original_filename).stem or "result"
        ext = "md" if type == "md" else "csv"
        return FileResponse(local, media_type="text/csv" if type == "csv" else "text/markdown", filename=f"{base}.{ext}")

    try:
        url = supabase_client.get_signed_download_url(path, bucket="results", expires_in=3600)
        return {"download_url": url}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f"다운로드 링크 생성 실패: {e}")
