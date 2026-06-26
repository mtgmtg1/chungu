#!/usr/bin/env python3
# [Flow: Step 1 (탭별 데이터 수집) -> Step 2 (openpyxl Workbook 생성) -> Step 3 (시트별 표 작성) -> Step 4 (파일 저장)]
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def _write_sheet(ws, rows: list[dict], columns: list[str]) -> None:
    """워크시트에 헤더와 행을 작성한다."""
    header = ["파일명", "유형", "위치/시간/페이지", "추출내용"] + columns
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append([row.get(c, "") for c in header])

    # 열 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = adjusted_width


def write_excel(tabs: dict[str, list[dict]], columns: list[str], out_path: Path) -> Path:
    """파일별 탭으로 구성된 Excel 파일을 작성한다.

    tabs: {sheet_name: [row_dict]}
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    first = True
    for sheet_name, rows in tabs.items():
        if first:
            ws = wb.active
            ws.title = sheet_name[:31]
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name[:31])
        _write_sheet(ws, rows, columns)
    wb.save(out_path)
    return out_path


def build_media_rows(filename: str, file_type: str, position: str, table_text: str) -> list[dict]:
    """미디어/이미지 처리 결과를 통일된 행 구조로 변환한다."""
    if not table_text:
        return [
            {
                "파일명": filename,
                "유형": file_type,
                "위치/시간/페이지": position,
                "추출내용": "",
            }
        ]
    rows: list[dict] = []
    lines = [ln.strip() for ln in table_text.strip().splitlines() if ln.strip().startswith("|")]
    if len(lines) < 2:
        # 표가 아니면 전체 텍스트를 한 행에 담는다
        return [
            {
                "파일명": filename,
                "유형": file_type,
                "위치/시간/페이지": position,
                "추출내용": table_text.strip(),
            }
        ]
    for line in lines[2:]:  # skip header + separator
        cells = [c.strip() for c in line.split("|")]
        cells = [c for c in cells if c]
        content = " | ".join(cells)
        rows.append(
            {
                "파일명": filename,
                "유형": file_type,
                "위치/시간/페이지": position,
                "추출내용": content,
            }
        )
    return rows


def build_pdf_rows(filename: str, page_tables: list[tuple[int, str]], columns: list[str]) -> list[dict]:
    """PDF vision/hybrid 결과를 통일된 행 구조로 변환한다."""
    rows: list[dict] = []
    for page_num, table_text in page_tables:
        from .merge import parse_markdown_table

        for record in parse_markdown_table(table_text):
            row: dict = {"파일명": filename, "유형": "pdf", "위치/시간/페이지": str(page_num), "추출내용": ""}
            for col in columns:
                row[col] = record.get(col, "")
            rows.append(row)
    return rows
