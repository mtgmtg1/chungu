#!/usr/bin/env python3
# [Flow: Step 1 (마크다운 입력) -> Step 2 (LLM에 구조화된 JSON 요청) -> Step 3 (JSON 파싱) -> Step 4 (openpyxl로 xlsx 작성)]
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .. import settings_store
from .ocr_client import call_text


SYSTEM_PROMPT = """You are a precise table conversion assistant. Your task is to convert a markdown table into a structured Excel-compatible JSON format.

Analyze the markdown table and return ONLY a JSON object in this exact format:
{
  "sheets": [
    {
      "name": "Sheet1",
      "headers": ["col1", "col2", ...],
      "rows": [
        ["val1", "val2", ...],
        ["val3", "val4", ...]
      ]
    }
  ]
}

Rules:
- Preserve all cell values exactly as they appear.
- If multiple markdown tables exist, create one sheet per table.
- If a table is logically split, merge them into one sheet if they share the same headers.
- Use empty string "" for missing cells.
- Return only the JSON object, no markdown code fences, no explanations.
"""


def _parse_llm_json(content: str) -> dict:
    """LLM 출력에서 JSON 블록을 추출하고 파싱한다."""
    content = content.strip()
    if content.startswith("```"):
        content = re.sub(r"```[a-zA-Z]*\n?|\n?```", "", content).strip()
    return json.loads(content)


def _write_xlsx(data: dict, out_path: Path) -> Path:
    """파싱된 JSON 데이터를 Excel 파일로 작성한다."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheets = data.get("sheets", [])
    if not sheets:
        sheets = [{"name": "Sheet1", "headers": [], "rows": []}]
    first = True
    for sheet in sheets:
        name = sheet.get("name", "Sheet1")[:31]
        headers = sheet.get("headers", []) or []
        rows = sheet.get("rows", []) or []
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)
        if headers:
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
    wb.save(out_path)
    return out_path


def convert_markdown_to_xlsx(
    markdown: str,
    out_path: Path,
    endpoint: str,
    model: str,
    api_key: str = "",
) -> Path:
    """마크다운 표를 LLM을 통해 분석하여 xlsx 파일로 변환한다."""
    if not markdown.strip():
        return _write_xlsx({"sheets": [{"name": "Sheet1", "headers": [], "rows": []}]}, out_path)
    prompt = f"{SYSTEM_PROMPT}\n\nMarkdown table to convert:\n\n{markdown}"
    content, _ = call_text(prompt, endpoint, model, api_key, max_tokens=4000)
    data = _parse_llm_json(content)
    return _write_xlsx(data, out_path)


def convert_markdown_to_xlsx_with_settings(
    markdown: str,
    out_path: Path,
    db,
    endpoint: str | None = None,
    model: str | None = None,
    api_key: str | None = None,
) -> Path:
    """설정/DB에서 LLM 설정을 읽어 마크다운을 xlsx로 변환한다."""
    ep = endpoint or settings_store.get_setting(db, "llm_endpoint") or ""
    md = model or settings_store.get_setting(db, "llm_model") or ""
    key = api_key if api_key is not None else settings_store.get_setting(db, "llm_api_key") or ""
    return convert_markdown_to_xlsx(markdown, out_path, ep, md, key)
