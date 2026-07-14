#!/usr/bin/env python3
# [Flow: Step 1 (원본 job 마크다운 로드) -> Step 2 (페이지 분할) -> Step 3 (첫 페이지에서 컬럼 구조 추출) -> Step 4 (페이지별 표 정리/재구성) -> Step 5 (xlsx 통합 저장) -> Step 6 (원래 job 업데이트)]
import json
import logging
import re
import tempfile
import traceback
from io import BytesIO
from pathlib import Path

from .. import settings_store
from ..config import settings
from ..core import office_converter, supabase_client
from ..core.ocr_client import call_text, call_vision, render_pdf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from ..db.models import Job, User
from ..db.session import SessionLocal

logger = logging.getLogger(__name__)

PAGE_MARKER_RE = re.compile(r"<!--\s*페이지\s*(\d+)\s*-->", re.IGNORECASE)
MAX_VISION_RETRIES = 3


def split_markdown_by_pages(markdown: str) -> list[tuple[int, str]]:
    """페이지 마커를 기준으로 마크다운을 분할한다."""
    matches = list(PAGE_MARKER_RE.finditer(markdown))
    if not matches:
        content = markdown.strip()
        if content:
            return [(1, content)]
        return []
    pages: list[tuple[int, str]] = []
    for idx, match in enumerate(matches):
        page_num = int(match.group(1))
        start = match.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(markdown)
        content = markdown[start:end].strip()
        if content:
            pages.append((page_num, content))
    return pages


def _get_markdown_content(job: Job) -> str:
    """편집된 마크다운이 있으면 사용하고, 없으면 원본 마크다운을 반환한다."""
    client = supabase_client.get_service_client()
    if job.result_edited_md_storage_path:
        data = client.storage.from_("results").download(job.result_edited_md_storage_path)
        return data.decode("utf-8")
    if job.result_md_storage_path:
        data = client.storage.from_("results").download(job.result_md_storage_path)
        return data.decode("utf-8")
    return ""


def _get_page_image_paths(
    job: Job,
    temp_dir: Path,
    page_range: list[int] | None = None,
) -> dict[int, Path]:
    """원본 파일에서 페이지별 이미지 경로를 생성한다. PDF는 렌더링, 이미지는 그대로 사용한다.

    [Flow: Step 1 (page_range가 주어지면 1-based 페이지 집합 생성) -> Step 2 (이미지 파일이면 page_range 필터링 후 복사)
          -> Step 3 (PDF면 page_range 범위만 렌더링) -> Step 4 (page_no → 이미지 경로 매핑 반환)]

    Args:
        job: Job 모델
        temp_dir: 임시 출력 디렉터리
        page_range: 1-based 페이지 번호 리스트. None이면 전체 페이지를 렌더링/복사한다.

    Returns:
        page_no(1-based) → 페이지 이미지 경로 매핑
    """
    image_paths: dict[int, Path] = {}
    page_set: set[int] = set(page_range) if page_range else set()
    files = job.extracted_files or []
    images = [
        (idx + 1, info)
        for idx, info in enumerate(files)
        if isinstance(info, dict) and info.get("type") == "image" and info.get("storage_path")
    ]
    if images:
        for page_num, info in images:
            if page_set and page_num not in page_set:
                continue
            try:
                data = supabase_client.get_service_client().storage.from_("pdfs").download(info["storage_path"])
                path = temp_dir / f"page-{page_num:03d}.png"
                path.write_bytes(data)
                image_paths[page_num] = path
            except Exception as e:
                logger.warning(f"[_get_page_image_paths] 이미지 다운로드 실패 page={page_num}: {e}")
        return image_paths

    if not job.pdf_storage_path:
        return image_paths

    try:
        input_bytes = supabase_client.download_pdf(job.pdf_storage_path).read()
        input_path = temp_dir / "input.pdf"
        input_path.write_bytes(input_bytes)
        if page_set:
            start_page = min(page_set)
            end_page = max(page_set)
            render_pdf(str(input_path), str(temp_dir), dpi=300, start=start_page, end=end_page)
            for p in sorted(temp_dir.glob("page-*.png")):
                try:
                    page_num = int(p.stem.split("-")[-1])
                    if page_num in page_set:
                        image_paths[page_num] = p
                except Exception:
                    pass
        else:
            render_pdf(str(input_path), str(temp_dir), dpi=300)
            for p in sorted(temp_dir.glob("page-*.png")):
                try:
                    page_num = int(p.stem.split("-")[-1])
                    image_paths[page_num] = p
                except Exception:
                    pass
    except Exception as e:
        logger.warning(f"[_get_page_image_paths] PDF 렌더링 실패: {e}")
    return image_paths


def _extract_column_structure(page_markdown: str, endpoint: str, model: str, api_key: str, user_language: str = "ko") -> dict:
    """첫 페이지에서 공통 컬럼 구조를 JSON으로 추출한다."""
    tables = office_converter._parse_markdown_tables(page_markdown)
    prompt = f"""Extract the common column structure for the entire document from the markdown table below. Return only JSON.

Format:
{{
  "columns": ["column1", "column2", ...],
  "description": "Briefly explain the meaning and data type of each column. Write in the user's configured language ({user_language})."
}}

Notes:
- If multiple tables appear consecutively, treat tables with the same header as one logical table and extract only the common columns applicable to the entire document.
- Ignore headers with empty or missing cells.
- Write column names concisely and clearly in the user's configured language ({user_language}).

Markdown:
{page_markdown}
"""
    try:
        content, _ = call_text(prompt, endpoint, model, api_key, max_tokens=2000)
        content = content.strip()
        if content.startswith("```"):
            content = re.sub(r"```[a-zA-Z]*\n?|\n?```", "", content).strip()
        data = json.loads(content)
        if isinstance(data, dict) and "columns" in data:
            return data
    except Exception as e:
        logger.warning(f"[_extract_column_structure] 컬럼 구조 추출 실패: {e}")
    # fallback: 첫 번째 표의 헤더 사용
    if tables:
        return {"columns": tables[0].get("headers", []), "description": ""}
    return {"columns": [], "description": ""}


def _parse_tables_safe(page_markdown: str) -> list[dict]:
    """마크다운 표를 파싱하고, 실패하면 빈 목록을 반환한다."""
    try:
        return office_converter._parse_markdown_tables(page_markdown)
    except Exception as e:
        logger.warning(f"[_parse_tables_safe] 표 파싱 실패: {e}")
        return []


def _normalize_page_with_llm(page_markdown: str, column_structure: dict, endpoint: str, model: str, api_key: str, user_language: str = "ko") -> dict:
    """LLM 텍스트로 페이지의 표를 공통 컬럼 구조에 맞춰 정리한다."""
    columns = column_structure.get("columns", [])
    prompt = f"""Analyze the markdown table below and organize it according to the common column structure. Return only JSON.

Common columns: {json.dumps(columns, ensure_ascii=False)}

Format:
{{
  "rows": [
    ["value1", "value2", ...],
    ["value3", "value4", ...]
  ],
  "valid": true,
  "reason": "Brief reason for whether the organization was successful. Write in the user's configured language ({user_language})."
}}

Rules:
- Each row must have the same length as the number of common columns. Fill missing cells with empty string "".
- Recover typos or missing cells from the original, but do not invent content. Use "" for uncertain values.
- Each column has a consistent data type (string/number/empty/account number/email/phone number/resident registration number/business registration number/amount/date, etc.). Align cells so the same formats go into the same column, and use format clues to determine the correct column when parsing errors cause column count or order mismatch.
- If the table is empty or the column structure is unknown, set valid=false and write the reason.

Markdown:
{page_markdown}
"""
    try:
        content, _ = call_text(prompt, endpoint, model, api_key, max_tokens=4000)
        content = content.strip()
        if content.startswith("```"):
            content = re.sub(r"```[a-zA-Z]*\n?|\n?```", "", content).strip()
        data = json.loads(content)
        if isinstance(data, dict):
            return data
    except Exception as e:
        logger.warning(f"[_normalize_page_with_llm] 페이지 정리 실패: {e}")
    return {"rows": [], "valid": False, "reason": "LLM 정리 실패"}


def _reconstruct_page_with_vision(image_path: Path, column_structure: dict, endpoint: str, model: str, api_key: str, page_markdown: str = "", user_language: str = "ko") -> dict:
    """vision LLM에 페이지 이미지를 전달하여 표를 재구성한다."""
    columns = column_structure.get("columns", [])
    prompt = f"""This image is a document page. Extract the table as JSON matching the common column structure below.

Common columns: {json.dumps(columns, ensure_ascii=False)}

Format:
{{
  "rows": [
    ["value1", "value2", ...],
    ["value3", "value4", ...]
  ],
  "valid": true,
  "reason": "Brief reason. Write in the user's configured language ({user_language})."
}}

Rules:
- Each row must have the same length as the number of common columns.
- Extract only what is visible in the image; do not guess content.
- Each column has a consistent data type (string/number/empty/account number/email/phone number/resident registration number/business registration number/amount/date, etc.). Align the same formats into the same column.
- If there is no table or it is unreadable, set valid=false and write the reason.
"""
    try:
        content, _ = call_vision(image_path, prompt, endpoint, model, api_key, max_tokens=4000, page_text=page_markdown)
        content = content.strip()
        if content.startswith("```"):
            content = re.sub(r"```[a-zA-Z]*\n?|\n?```", "", content).strip()
        data = json.loads(content)
        if isinstance(data, dict):
            return data
    except Exception as e:
        logger.warning(f"[_reconstruct_page_with_vision] vision 재구성 실패: {e}")
    return {"rows": [], "valid": False, "reason": "vision 재구성 실패"}


def _evaluate_reconstruction(image_path: Path, rows: list[list[str]], column_structure: dict, endpoint: str, model: str, api_key: str, user_language: str = "ko") -> tuple[bool, str]:
    """vision LLM이 재구성한 결과를 다시 평가하여 재시도 여부를 결정한다."""
    columns = column_structure.get("columns", [])
    if not rows:
        return False, "rows are empty"
    prompt = f"""Compare this image with the extracted table below and evaluate whether the table was accurately extracted.

Common columns: {json.dumps(columns, ensure_ascii=False)}
Number of extracted rows: {len(rows)}

Format:
{{
  "ok": true,
  "reason": "'Accurate' if correct, or specific issues if there are problems. Write in the user's configured language ({user_language})."
}}

If ok is false, a retry is needed."""
    try:
        content, _ = call_vision(image_path, prompt, endpoint, model, api_key, max_tokens=2000)
        content = content.strip()
        if content.startswith("```"):
            content = re.sub(r"```[a-zA-Z]*\n?|\n?```", "", content).strip()
        data = json.loads(content)
        if isinstance(data, dict):
            return bool(data.get("ok", False)), str(data.get("reason", ""))
    except Exception as e:
        logger.warning(f"[_evaluate_reconstruction] 평가 실패: {e}")
    return False, "평가 실패"


def _pad_row(row: list[str], col_count: int) -> list[str]:
    """행을 지정한 컬럼 수에 맞춘다."""
    return row + [""] * (col_count - len(row)) if len(row) < col_count else row[:col_count]


def _merge_page_tables(tables: list[dict], recovery_notes: list[dict]) -> list[dict]:
    """페이지별 표를 하나로 통합하며, 동일 헤더는 연결하고 다른 헤더는 빈 행으로 구분한다."""
    if not tables:
        return []
    merged: list[dict] = [tables[0]]
    for table in tables[1:]:
        last = merged[-1]
        if table["headers"] == last["headers"]:
            last["rows"].extend(table["rows"])
        else:
            merged.append(table)
    return office_converter._normalize_rows(merged)


def _write_xlsx_advanced(tables: list[dict], recovery_notes: list[dict], out_path: Path) -> Path:
    """통합된 표와 복구 노트를 하나의 xlsx 파일로 작성한다."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tables"

    row = 1
    for idx, table in enumerate(tables):
        if idx > 0:
            row += 1
        headers = table["headers"]
        for col_idx, h in enumerate(headers):
            cell = ws.cell(row, col_idx + 1, h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1
        for table_row in table["rows"]:
            for col_idx, val in enumerate(table_row):
                ws.cell(row, col_idx + 1, val)
            row += 1

    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(out_path)
    return out_path


def run(parent_job_id: str) -> dict:
    """고급 변환을 실행하고 원래 job의 상태를 업데이트한다."""
    db = SessionLocal()
    job = db.get(Job, parent_job_id)
    if job is None:
        db.close()
        return {"error": "job not found"}

    user_language = "ko"
    if job.user_id:
        try:
            user = db.get(User, job.user_id)
            if user and user.language:
                user_language = user.language
        except Exception:
            pass

    endpoint = job.endpoint or settings_store.get_setting(db, "llm_endpoint") or settings.default_llm_endpoint
    model = job.model or settings_store.get_setting(db, "llm_model") or settings.default_llm_model
    api_key = settings_store.get_setting(db, "llm_api_key") or ""

    units = job.total_pages if job.total_pages else (job.total_files or 1)
    cost = units * 3

    try:
        markdown = _get_markdown_content(job)
        pages = split_markdown_by_pages(markdown)
        if not pages:
            raise ValueError("No markdown pages to convert")

        with tempfile.TemporaryDirectory() as tmpdir:
            temp_dir = Path(tmpdir)
            image_paths = _get_page_image_paths(job, temp_dir)

            # 첫 페이지에서 컬럼 구조 추출
            first_page_num, first_page_content = pages[0]
            column_structure = _extract_column_structure(first_page_content, endpoint, model, api_key, user_language=user_language)
            if not column_structure.get("columns"):
                # 첫 페이지에 표가 없으면 다른 페이지에서 찾기
                for page_num, page_content in pages[1:]:
                    column_structure = _extract_column_structure(page_content, endpoint, model, api_key, user_language=user_language)
                    if column_structure.get("columns"):
                        break
            columns = column_structure.get("columns", [])
            if not columns:
                raise ValueError("Could not extract table column structure from document")

            all_tables: list[dict] = []
            recovery_notes: list[dict] = []
            any_valid = False

            for page_num, page_content in pages:
                page_tables = _parse_tables_safe(page_content)
                normalized_rows: list[list[str]] = []
                valid = False
                reason = ""
                used_vision = False

                if page_tables:
                    # 텍스트 기반 정리
                    result = _normalize_page_with_llm(page_content, column_structure, endpoint, model, api_key, user_language=user_language)
                    if result.get("valid") and result.get("rows"):
                        normalized_rows = [_pad_row(r, len(columns)) for r in result["rows"]]
                        valid = True
                        reason = result.get("reason", "")

                if not valid and page_num in image_paths:
                    # vision LLM으로 재구성 (최대 3번 재시도)
                    used_vision = True
                    eval_reason = ""
                    for attempt in range(1, MAX_VISION_RETRIES + 1):
                        vision_result = _reconstruct_page_with_vision(
                            image_paths[page_num], column_structure, endpoint, model, api_key, page_content, user_language=user_language
                        )
                        rows = [_pad_row(r, len(columns)) for r in vision_result.get("rows", [])]
                        ok, eval_reason = _evaluate_reconstruction(
                            image_paths[page_num], rows, column_structure, endpoint, model, api_key, user_language=user_language
                        )
                        if ok and rows:
                            normalized_rows = rows
                            valid = True
                            reason = f"vision 재구성 성공 (시도 {attempt})"
                            break

                    if not valid and normalized_rows:
                        # 재구성은 됐지만 평가가 실패한 경우, 복구용 셀로 사용
                        recovery_notes.append({
                            "page": page_num,
                            "reason": f"vision 재구성 후 평가 실패: {eval_reason}",
                            "cell": "해당 페이지 전체",
                        })
                        # 전체 행을 복구용 셀로 표시하지 않고, 기존 rows를 사용
                        valid = True

                if not valid:
                    # 복구용 셀로 표 구조 유지
                    normalized_rows = [["[복구 실패: 페이지 표 추출 불가]"] + [""] * (len(columns) - 1)]
                    recovery_notes.append({
                        "page": page_num,
                        "reason": "텍스트/vision 모두 실패하여 복구용 셀 삽입",
                        "cell": f"A{len(all_tables) + 1}",
                    })

                if normalized_rows:
                    any_valid = True
                    all_tables.append({"headers": columns, "rows": normalized_rows})

            if not any_valid:
                raise ValueError("Could not extract tables from any page; cannot generate Excel")

            merged_tables = _merge_page_tables(all_tables, recovery_notes)
            out_path = temp_dir / "result_advanced.xlsx"
            _write_xlsx_advanced(merged_tables, recovery_notes, out_path)
            storage_path = supabase_client.upload_office_result(job.id, out_path, "xlsx")

            job.result_xlsx_advanced_storage_path = storage_path
            job.xlsx_advanced_converted = True
            job.xlsx_advanced_status = "done"
            job.xlsx_advanced_recovery_notes = recovery_notes
            job.xlsx_advanced_refundable = False
            db.commit()
            return {"job_id": parent_job_id, "status": "done", "recovery_notes": len(recovery_notes)}

    except Exception as e:
        logger.exception(f"[convert_xlsx_advanced] {parent_job_id} 변환 실패: {e}")
        tb = traceback.format_exc()
        job.xlsx_advanced_status = "error"
        job.xlsx_advanced_refundable = True
        job.xlsx_advanced_recovery_notes = job.xlsx_advanced_recovery_notes + [{
            "page": 0,
            "reason": f"전체 변환 실패: {e}\n{tb}",
            "cell": "-",
        }]
        db.commit()
        # 환불은 프론트엔드에서 재시도/환불 선택 시 수행
        return {"job_id": parent_job_id, "status": "error", "error": str(e)}
    finally:
        db.close()
